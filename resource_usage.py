################################################################################
#Generate a  memory usage summary excel file  from a linker map file  
#Usage: python resource_usage.py input_file [output_file]
#Author :Mahesh C
#
###############################################################################


import sys

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.cell import Cell
except :
    print "Unable load openpyxl libs, make sure it is avaialble in PYTHON_PATH"



#Change this parameters to modify the exceution behaviour  

CONFIG={'ram_range':['0x40000000','0x49000000'],  #Consider only this ram range
        'rom_range':['0x00000000','0x40000000'],  #Consider only this rom  range
        'identifiers':["\\obj\\","\\lib\\"], # identify the section based on this identifiers
        'ram_address_msd':'4', #RAM Memory will start from this msd'
        'default_output_file':'out.xls',   # Result file
        'start_reading_from': "Link Editor Memory Map", #Mem usage sections in the input_file will start after this line
        'avoid_lines_with':'.debug' # skip the lines with this keyword
        }


class Map():
    def __init__(self,_line,_identifier):
        self.identifier=_identifier
        valid_data=filter(None,_line.split(" "))
        self.section=valid_data[0].replace("\t","")
        self.address=valid_data[1].split("\t")[0]
        self.size=int(valid_data[1].split("\t")[1],16)
        self.file_name=valid_data[-1].split(_identifier)[1].replace("\r","").replace("\n","")
        
            
        
class Parser():
    def __init__(self,_file):
        self.in_file=open(_file,"r")
        data=self.in_file.readlines()
        start_map=0
        self.maps=[]
        for line in data:
            if line.find(CONFIG['start_reading_from'])>-1:
                start_map=1
            if line.find(CONFIG['avoid_lines_with'])>-1:
                start_map=0
            if start_map==1:
                for identifier in CONFIG['identifiers']:
                    if line.find(identifier)>-1:
                        c_map=Map(line,identifier)
                        self.maps.append(c_map)
                        break
    def getMaps(self):
        return self.maps
               
        
class XlsWriter():
    def __init__(self,_file=CONFIG['default_output_file'],maps=None,sheet_name=None):
        self.maps=maps
        self.output_file=_file
        if sheet_name is None:
            sheet_name='RAWDATA'
        if maps is None:
            print "Empty maps"
        try: #Existing output file
            self.wb = load_workbook(_file)
            self.ws_summary=self.wb.get_sheet_by_name("Summary")
            
        except:#new file
            self.wb = Workbook()
            self.ws_summary = self.wb.active
            self.ws_summary.title="Summary"
        self.ws_data = self.wb.create_sheet()
        self.ws_data.title=sheet_name
        self.data={}
        for identifier in CONFIG['identifiers']:
            self.data[identifier]={'rom_usage':0,'ram_usage':0,'rom_maps':[],'ram_maps':[]}
    def  sortData(self):
        for _map in self.maps:
            
            if _map.address[0]==CONFIG['ram_address_msd']:
                if int(CONFIG['ram_range'][0],16)<=int(_map.address,16) and int(CONFIG['ram_range'][1],16)>=int(_map.address,16):
                    self.data[_map.identifier]['ram_maps'].append(_map)
                    self.data[_map.identifier]['ram_usage']=self.data[_map.identifier]['ram_usage']+_map.size
            else:
                if int(CONFIG['rom_range'][0],16)<=int(_map.address,16) and int(CONFIG['rom_range'][1],16)>=int(_map.address,16):
                    self.data[_map.identifier]['rom_maps'].append(_map)
                    self.data[_map.identifier]['rom_usage']=self.data[_map.identifier]['rom_usage']+_map.size
    def writeRaw(self,maps,start_from,sec_title,mem_title):
        _i=start_from
        self.ws_data.cell(row = _i, column = 1).value=sec_title
        self.ws_data.cell(row = _i, column = 2).value=mem_title
        self.ws_data.cell(row = _i+1, column = 1).value='Section'
        self.ws_data.cell(row = _i+1, column = 2).value='Address'
        self.ws_data.cell(row = _i+1, column = 3).value='Size'
        self.ws_data.cell(row = _i+1, column = 4).value='FileName'
        _i=_i+2
        for _map in maps:
            d = self.ws_data.cell(row = _i, column = 1)
            d.value=_map.section
            d = self.ws_data.cell(row = _i, column = 2)
            d.value=_map.address
            d = self.ws_data.cell(row = _i, column = 3)
            d.value=_map.size
            d = self.ws_data.cell(row = _i, column = 4)
            d.value=_map.file_name
            _i=_i+1
        return _i
    

    
    
    def write(self):
        s_r=self.ws_summary.get_highest_row()+3
        s_c=self.ws_summary.get_highest_column()
     
        self.ws_summary.cell(row = s_r-1, column = 5).value=input_file


        self.ws_summary.cell(row = s_r, column = 5).value="Resource"
        
        c_i=1
        for identifier in CONFIG['identifiers']:
            self.ws_summary.cell(row = s_r, column = 5+c_i).value=identifier+'(bytes)'
            c_i=c_i+1
        c_i=5+c_i    
        self.ws_summary.cell(row = s_r, column = c_i).value="Total bytes"
        self.ws_summary.cell(row = s_r, column = c_i+1).value="Total in KB"
        

        self.ws_summary.cell(row = s_r+1, column = 5).value="RAM"
        c_i=1
        for identifier in CONFIG['identifiers']:
            self.ws_summary.cell(row = s_r+1, column = 5+c_i).value=self.data[identifier]['ram_usage']
            c_i=c_i+1
        c_i=5+c_i    
        total_ram=0
        for identifier in CONFIG['identifiers']:
            total_ram=total_ram+self.data[identifier]['ram_usage']
        self.ws_summary.cell(row = s_r+1, column = c_i).value=total_ram
        self.ws_summary.cell(row = s_r+1, column = c_i+1).value=total_ram/1024

        self.ws_summary.cell(row = s_r+2, column = 5).value="ROM"
        c_i=1
        for identifier in CONFIG['identifiers']:
            self.ws_summary.cell(row = s_r+2, column = 5+c_i).value=self.data[identifier]['rom_usage']
            c_i=c_i+1
        c_i=5+c_i    
        total_rom=0
        for identifier in CONFIG['identifiers']:
            total_rom=total_rom+self.data[identifier]['rom_usage']
        self.ws_summary.cell(row = s_r+2, column = c_i).value=total_rom
        self.ws_summary.cell(row = s_r+2, column = c_i+1).value=total_rom/1024
        
        
        #TODO move to dynamic indexing of cols
        self.ws_summary.column_dimensions['E'].width = 15
        self.ws_summary.column_dimensions['F'].width = 15
        self.ws_summary.column_dimensions['G'].width = 15
        self.ws_summary.column_dimensions['H'].width = 15
        self.ws_summary.column_dimensions['I'].width = 15
        
        s_i=0
        for identifier in CONFIG['identifiers']:
            sorted_obj = sorted(self.data[identifier]['rom_maps'], key=lambda x: x.size, reverse=True)
            s_i=self.writeRaw(sorted_obj,s_i+1,identifier,'ROM')
            sorted_obj = sorted(self.data[identifier]['ram_maps'], key=lambda x: x.size, reverse=True)
            s_i=self.writeRaw(sorted_obj,s_i+1,identifier,'RAM')
        
        
        
        self.wb.save(self.output_file)

if __name__=='__main__':
    if len(sys.argv)<2:
        print "Usage: python resource_usage.py input_file [output_file]"
    else:
        input_file=sys.argv[1]
        ldfile_parser=Parser(input_file)
        if len(sys.argv)>2:
            output_file=sys.argv[2]
            writer=XlsWriter(output_file,ldfile_parser.getMaps(),sheet_name=input_file)
        else:
             writer=XlsWriter(maps=ldfile_parser.getMaps(),sheet_name=input_file)
        writer.sortData()
        writer.write()







